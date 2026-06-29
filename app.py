import streamlit as st
import os
import psycopg2
import pandas as pd
from urllib.parse import urlparse, urlunparse
from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook, load_workbook
import unicodedata
import hashlib
import altair as alt
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer



# =========================
# CONFIGURAÇÃO GERAL
# =========================

st.set_page_config(
    page_title="Lista de Presença RH",
    layout="wide"
)

TEMPLATE_EXCEL = "Lista de presença.xlsx"
FILIAL_PADRAO = "MG CGE"

STATUS_PRESENCA = [
    "Presente",
    "Falta",
    "Atestado",
    "Folga",
    "Folga Dominical",
    "DSR",
    "Férias",
    "Afastamento",
    "Desligado",
    "Integração",
    "Feriado"
]

STATUS_ABSENTEISMO = [
    "Falta",
    "Atestado",
    "Afastamento"
]

COLUNAS_IMPORTACAO = [
    "matricula",
    "nome",
    "jornada_trabalho",
    "cargo",
    "setor",
    "logins_jms",
    "gestor_responsavel",
    "folga_dominical",
    "genero",
    "ativo"
]

PERFIS_USUARIO = [
    "Gestor",
    "Operação"
]

IDIOMAS_DISPONIVEIS = {
    "Português (PT-BR)": "pt",
    "中文简体": "zh"
}

TRADUCOES = {
    "pt": {},
    "zh": {
        "Lista de Presença RH": "人力资源考勤表",
        "Usuário": "用户",
        "Senha": "密码",
        "Entrar": "登录",
        "Usuário ou senha inválidos.": "用户名或密码无效。",
        "Idioma": "语言",
        "Painel do Gestor - RH": "人力资源管理面板",
        "Operação - Presença RH": "运营 - 人力资源考勤",
        "Dashboard": "仪表板",
        "Importar Dados": "导入数据",
        "Cadastrar / Alterar Pessoas": "新增 / 修改人员",
        "Colaboradores": "员工",
        "Escala Folga Dominical": "周日休息排班",
        "Exportar Excel": "导出 Excel",
        "Acessos": "访问权限",
        "Lançamento": "考勤录入",
        "Histórico": "历史记录",
        "Lançamento de Presença": "考勤录入",
        "Histórico de Presença": "考勤历史",
        "Dashboard gerencial": "管理仪表板",
        "Importar dados de colaboradores": "导入员工数据",
        "Cadastro e alteração de colaboradores": "员工新增与修改",
        "Base de colaboradores": "员工数据库",
        "Escala de Folga Dominical": "周日休息排班",
        "Exportar Excel": "导出 Excel",
        "Gestão de acessos": "访问权限管理",
        "Data": "日期",
        "Data inicial": "开始日期",
        "Data final": "结束日期",
        "Responsável": "负责人",
        "Todos": "全部",
        "Matrícula": "工号",
        "Nome": "姓名",
        "Cargo": "职位",
        "Status": "状态",
        "Observação": "备注",
        "Selecione": "请选择",
        "Salvar novos lançamentos": "保存新记录",
        "Salvar alterações": "保存修改",
        "Alterar lançamentos desta data": "修改当天记录",
        "Sair": "退出",
        "Perfil": "角色",
        "Filial": "分部",
        "Headcount": "在岗人数",
        "Turnover": "人员流动率",
        "Desligados": "离职人数",
        "Registros": "记录数",
        "Presentes": "出勤",
        "Faltas": "缺勤",
        "Ausências ABS": "缺勤ABS",
        "Baixar histórico em Excel": "下载历史 Excel",
        "Baixar modelo de importação": "下载导入模板",
        "Arquivo Excel": "Excel 文件",
        "Aba da planilha": "工作表",
        "Linha onde está o cabeçalho": "表头所在行",
        "Importar dados para o sistema": "导入到系统",
        "Salvar alterações da base": "保存数据库修改",
        "Cadastrar novo": "新增",
        "Alterar cadastro": "修改资料",
        "Cadastrar": "新增",
        "Ativo": "启用",
        "Gênero": "性别",
        "Folga Dominical": "周日休息",
        "Jornada de Trabalho": "工作班次",
        "Setor": "部门",
        "Logins - JMS": "JMS 登录",
        "Baixar Excel": "下载 Excel",
        "Baixar PDF paisagem": "下载横向 PDF",
        "Gravar escala no histórico": "保存排班到历史",
        "Ano": "年份",
        "Mês": "月份",
        "Criar novo acesso": "创建新访问",
        "Criar acesso": "创建访问",
        "Alterar acesso existente": "修改现有访问",
        "Nova senha": "新密码",
        "Salvar alteração do acesso": "保存访问修改",
        "Usuários cadastrados": "已注册用户"
    }
}

def obter_codigo_idioma():
    return st.session_state.get("idioma", "pt")

def t(texto):
    return TRADUCOES.get(obter_codigo_idioma(), {}).get(texto, texto)

def seletor_idioma(chave="idioma_selector"):
    idioma_atual = st.session_state.get("idioma", "pt")
    opcoes = list(IDIOMAS_DISPONIVEIS.keys())
    indice = 0
    for i, rotulo in enumerate(opcoes):
        if IDIOMAS_DISPONIVEIS[rotulo] == idioma_atual:
            indice = i
            break
    selecionado = st.selectbox(t("Idioma"), opcoes, index=indice, key=chave)
    st.session_state["idioma"] = IDIOMAS_DISPONIVEIS[selecionado]


# =========================
# UTILITÁRIOS
# =========================

def normalizar_texto(texto):
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.replace("_", " ").replace("-", " ")
    texto = " ".join(texto.split())
    return texto


def hash_senha(senha):
    return hashlib.sha256(str(senha).encode("utf-8")).hexdigest()


def converter_ativo(valor):
    if isinstance(valor, bool):
        return 1 if valor else 0

    texto = str(valor).strip().lower()

    if texto in ["", "1", "sim", "s", "ativo", "true", "verdadeiro"]:
        return 1

    if texto in ["0", "não", "nao", "n", "inativo", "false", "falso", "desligado"]:
        return 0

    return 1


def primeiro_valor(df, coluna, padrao=""):
    if coluna not in df.columns or df.empty:
        return padrao

    valores = df[coluna].dropna()

    if valores.empty:
        return padrao

    return str(valores.iloc[0])


def opcoes_unicas(df, coluna, texto_todos):
    if coluna not in df.columns or df.empty:
        return [texto_todos]

    opcoes = sorted([
        str(item).strip()
        for item in df[coluna].dropna().unique()
        if str(item).strip()
    ])

    return [texto_todos] + opcoes


def filtrar_dataframe_colaborador(df, termo):
    if df.empty or not termo.strip():
        return df

    termo_normalizado = normalizar_texto(termo)

    mascara = (
        df["nome"].astype(str).apply(normalizar_texto).str.contains(termo_normalizado, na=False) |
        df["matricula"].astype(str).apply(normalizar_texto).str.contains(termo_normalizado, na=False) |
        df["cargo"].astype(str).apply(normalizar_texto).str.contains(termo_normalizado, na=False) |
        df["setor"].astype(str).apply(normalizar_texto).str.contains(termo_normalizado, na=False) |
        df["gestor_responsavel"].astype(str).apply(normalizar_texto).str.contains(termo_normalizado, na=False)
    )

    return df[mascara].copy()


# =========================
# BANCO DE DADOS - SUPABASE POSTGRES
# =========================

DATABASE_URL_SECRET_NAME = "SUPABASE_DB_URL"


def limpar_database_url(url):
    """Normaliza erros comuns ao colar a connection string nos Secrets do Streamlit."""
    if not url:
        return None

    url = str(url).strip().strip('"').strip("'").strip()

    # Corrige casos como: SUPABASE_DB_URL = "SUPABASE_DB_URL = "postgresql://...""
    candidatos = [
        pos for pos in [
            url.find("postgresql://"),
            url.find("postgres://")
        ]
        if pos >= 0
    ]

    if candidatos:
        url = url[min(candidatos):]

    return url.strip().strip('"').strip("'").strip()


def obter_database_url():
    """Busca a connection string do Supabase em secrets.toml ou variável de ambiente."""
    url = os.getenv(DATABASE_URL_SECRET_NAME) or os.getenv("DATABASE_URL")

    if not url:
        try:
            url = st.secrets.get(DATABASE_URL_SECRET_NAME) or st.secrets.get("DATABASE_URL")
        except Exception:
            url = None

    return limpar_database_url(url)


def mascarar_database_url(database_url):
    """Mostra a URL sem expor a senha."""
    try:
        partes = urlparse(database_url)
        host_port_db = partes.netloc.split("@")[-1]
        usuario = partes.username or "usuario"
        netloc_seguro = f"{usuario}:********@{host_port_db}"
        return urlunparse(partes._replace(netloc=netloc_seguro))
    except Exception:
        return "postgresql://usuario:********@host:porta/database"


def validar_database_url(database_url):
    problemas = []

    if not database_url:
        problemas.append("A variável SUPABASE_DB_URL não foi encontrada.")
        return problemas

    if "[YOUR-PASSWORD]" in database_url:
        problemas.append("A senha ainda está como [YOUR-PASSWORD]. Substitua pela senha real do banco.")

    if not (database_url.startswith("postgresql://") or database_url.startswith("postgres://")):
        problemas.append("A URL precisa começar com postgresql:// ou postgres://.")

    if database_url.count("SUPABASE_DB_URL") > 0:
        problemas.append("O valor do Secret não deve conter o texto SUPABASE_DB_URL dentro da URL.")

    try:
        partes = urlparse(database_url)

        if not partes.username:
            problemas.append("Usuário do banco não identificado na URL.")

        if not partes.password:
            problemas.append("Senha do banco não identificada na URL.")

        if not partes.hostname:
            problemas.append("Host do Supabase não identificado na URL.")

        try:
            porta = partes.port
        except ValueError:
            porta = None
            problemas.append("Porta inválida. Verifique se caracteres especiais da senha foram codificados.")

        if porta not in [5432, 6543]:
            problemas.append("Porta incomum para Supabase. Normalmente é 5432 ou 6543.")

        if partes.fragment:
            problemas.append("A URL contém # não codificado. Na senha, use %23 no lugar de #.")

        if partes.query:
            problemas.append("A URL contém ? ou & não codificado. Na senha, use %3F para ? e %26 para &.")

    except Exception:
        problemas.append("Não foi possível interpretar a URL do banco. Revise a sintaxe do Secret.")

    return problemas


def exibir_erro_conexao_supabase(database_url, erro=None):
    st.error("Não foi possível conectar ao Supabase Postgres.")

    st.markdown("""
**Plano de correção:**

1. Confirme se a senha usada é a senha do banco, não a senha da sua conta Supabase.
2. No Supabase, copie novamente a connection string em **Connect > Transaction pooler**.
3. Substitua `[YOUR-PASSWORD]` pela senha real do banco.
4. Se a senha tiver caracteres especiais, codifique-os na URL.
5. Salve o Secret no Streamlit e use **Reboot app**.
""")

    st.caption("Connection string detectada, com senha mascarada:")
    st.code(mascarar_database_url(database_url), language="text")

    if erro:
        detalhe = str(erro).strip()
        detalhe_lower = detalhe.lower()

        if "password authentication failed" in detalhe_lower:
            st.warning("Diagnóstico provável: senha incorreta ou senha não codificada corretamente na URL.")
        elif "could not translate host name" in detalhe_lower:
            st.warning("Diagnóstico provável: host copiado incorretamente da connection string.")
        elif "timeout" in detalhe_lower or "timed out" in detalhe_lower:
            st.warning("Diagnóstico provável: timeout de rede, projeto Supabase pausado ou endpoint/porta incorretos.")
        elif "ssl" in detalhe_lower:
            st.warning("Diagnóstico provável: problema de SSL. O app está usando sslmode=require.")
        else:
            st.warning("Diagnóstico provável: credencial, host, porta ou projeto Supabase inativo.")

        st.caption("Detalhe técnico seguro, sem senha:")
        st.code(detalhe.replace(database_url, mascarar_database_url(database_url)), language="text")

    st.stop()


def conectar():
    database_url = obter_database_url()

    if not database_url:
        st.error(
            "Connection string do Supabase não configurada. "
            "Cadastre SUPABASE_DB_URL nos Secrets do Streamlit."
        )
        st.stop()

    problemas = validar_database_url(database_url)

    if problemas:
        st.error("A variável SUPABASE_DB_URL foi encontrada, mas está com problema de configuração.")
        for problema in problemas:
            st.write(f"- {problema}")
        st.caption("Formato esperado:")
        st.code(
            'SUPABASE_DB_URL = "postgresql://postgres.PROJECT_REF:SENHA_DO_BANCO@aws-REGION.pooler.supabase.com:6543/postgres"',
            language="toml"
        )
        st.stop()

    try:
        return psycopg2.connect(
            database_url,
            sslmode="require",
            connect_timeout=15
        )
    except psycopg2.OperationalError as erro:
        exibir_erro_conexao_supabase(database_url, erro)


def obter_colunas(cursor, tabela):
    cursor.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = %s
    """, (tabela,))

    return [item[0] for item in cursor.fetchall()]


def garantir_coluna(cursor, tabela, coluna, tipo_sql):
    colunas = obter_colunas(cursor, tabela)

    if coluna not in colunas:
        cursor.execute(f"ALTER TABLE public.{tabela} ADD COLUMN {coluna} {tipo_sql}")


def criar_tabelas():
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id BIGSERIAL PRIMARY KEY,
            usuario TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            perfil TEXT NOT NULL,
            ativo INTEGER DEFAULT 1,
            criado_em TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS colaboradores (
            id BIGSERIAL PRIMARY KEY,
            matricula TEXT,
            nome TEXT NOT NULL,
            jornada_trabalho TEXT,
            cargo TEXT,
            setor TEXT,
            filial TEXT DEFAULT 'MG CGE',
            logins_jms TEXT,
            gestor_responsavel TEXT,
            folga_dominical TEXT,
            genero TEXT,
            ativo INTEGER DEFAULT 1
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS presencas (
            id BIGSERIAL PRIMARY KEY,
            colaborador_id BIGINT NOT NULL REFERENCES colaboradores(id) ON DELETE CASCADE,
            data DATE NOT NULL,
            status TEXT NOT NULL,
            observacao TEXT,
            criado_em TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
            atualizado_em TIMESTAMPTZ
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS escalas_folga_dominical (
            id BIGSERIAL PRIMARY KEY,
            colaborador_id BIGINT NOT NULL REFERENCES colaboradores(id) ON DELETE CASCADE,
            data_domingo DATE NOT NULL,
            status TEXT DEFAULT 'Folga Dominical',
            motivo_regra TEXT,
            criado_em TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
            atualizado_em TIMESTAMPTZ
        )
    """)

    garantir_coluna(cursor, "colaboradores", "matricula", "TEXT")
    garantir_coluna(cursor, "colaboradores", "jornada_trabalho", "TEXT")
    garantir_coluna(cursor, "colaboradores", "cargo", "TEXT")
    garantir_coluna(cursor, "colaboradores", "setor", "TEXT")
    garantir_coluna(cursor, "colaboradores", "filial", "TEXT DEFAULT 'MG CGE'")
    garantir_coluna(cursor, "colaboradores", "logins_jms", "TEXT")
    garantir_coluna(cursor, "colaboradores", "gestor_responsavel", "TEXT")
    garantir_coluna(cursor, "colaboradores", "folga_dominical", "TEXT")
    garantir_coluna(cursor, "colaboradores", "genero", "TEXT")
    garantir_coluna(cursor, "colaboradores", "ativo", "INTEGER DEFAULT 1")

    garantir_coluna(cursor, "presencas", "status", "TEXT")
    garantir_coluna(cursor, "presencas", "observacao", "TEXT")
    garantir_coluna(cursor, "presencas", "criado_em", "TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP")
    garantir_coluna(cursor, "presencas", "atualizado_em", "TIMESTAMPTZ")

    garantir_coluna(cursor, "escalas_folga_dominical", "status", "TEXT DEFAULT 'Folga Dominical'")
    garantir_coluna(cursor, "escalas_folga_dominical", "motivo_regra", "TEXT")
    garantir_coluna(cursor, "escalas_folga_dominical", "criado_em", "TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP")
    garantir_coluna(cursor, "escalas_folga_dominical", "atualizado_em", "TIMESTAMPTZ")

    cursor.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_escala_folga_dominical_colaborador_data
        ON escalas_folga_dominical (colaborador_id, data_domingo)
    """)

    cursor.execute("""
        INSERT INTO usuarios (
            usuario,
            senha_hash,
            perfil,
            ativo
        )
        VALUES (%s, %s, %s, 1)
        ON CONFLICT (usuario) DO NOTHING
    """, (
        "gestor",
        hash_senha("1234"),
        "Gestor"
    ))

    cursor.execute("""
        INSERT INTO usuarios (
            usuario,
            senha_hash,
            perfil,
            ativo
        )
        VALUES (%s, %s, %s, 1)
        ON CONFLICT (usuario) DO NOTHING
    """, (
        "operacao",
        hash_senha("1234"),
        "Operação"
    ))

    cursor.execute("""
        UPDATE colaboradores
        SET filial = %s
        WHERE filial IS NULL OR TRIM(filial) = ''
    """, (FILIAL_PADRAO,))

    cursor.execute("""
        DELETE FROM presencas
        WHERE id NOT IN (
            SELECT MAX(id)
            FROM presencas
            GROUP BY colaborador_id, data
        )
    """)

    cursor.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_presenca_colaborador_data
        ON presencas (colaborador_id, data)
    """)

    conn.commit()
    conn.close()


# =========================
# USUÁRIOS / ACESSOS
# =========================

def autenticar_usuario(usuario, senha):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT usuario, perfil
        FROM usuarios
        WHERE usuario = %s
          AND senha_hash = %s
          AND ativo = 1
        LIMIT 1
    """, (
        str(usuario).strip(),
        hash_senha(senha)
    ))

    resultado = cursor.fetchone()
    conn.close()

    if resultado:
        return {
            "usuario": resultado[0],
            "perfil": resultado[1]
        }

    return None


def listar_usuarios():
    conn = conectar()

    df = pd.read_sql_query("""
        SELECT
            id,
            usuario,
            perfil,
            ativo,
            criado_em
        FROM usuarios
        ORDER BY usuario
    """, conn)

    conn.close()

    if not df.empty:
        df["ativo"] = df["ativo"].apply(lambda x: bool(int(x)))

    return df


def criar_usuario(usuario, senha, perfil):
    usuario = str(usuario).strip()

    if not usuario or not senha or perfil not in PERFIS_USUARIO:
        return {
            "sucesso": False,
            "mensagem": "Usuário, senha e perfil são obrigatórios."
        }

    conn = conectar()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO usuarios (
                usuario,
                senha_hash,
                perfil,
                ativo
            )
            VALUES (%s, %s, %s, 1)
        """, (
            usuario,
            hash_senha(senha),
            perfil
        ))

        conn.commit()
        conn.close()

        return {
            "sucesso": True,
            "mensagem": "Acesso criado com sucesso."
        }

    except psycopg2.errors.UniqueViolation:
        conn.close()

        return {
            "sucesso": False,
            "mensagem": "Esse usuário já existe."
        }


def atualizar_usuario(usuario_id, perfil, ativo, nova_senha=None):
    usuario_id = int(usuario_id)
    conn = conectar()
    cursor = conn.cursor()

    if nova_senha:
        cursor.execute("""
            UPDATE usuarios
            SET
                perfil = %s,
                ativo = %s,
                senha_hash = %s
            WHERE id = %s
        """, (
            perfil,
            converter_ativo(ativo),
            hash_senha(nova_senha),
            usuario_id
        ))
    else:
        cursor.execute("""
            UPDATE usuarios
            SET
                perfil = %s,
                ativo = %s
            WHERE id = %s
        """, (
            perfil,
            converter_ativo(ativo),
            usuario_id
        ))

    conn.commit()
    conn.close()


# =========================
# COLABORADORES
# =========================

def cadastrar_colaborador(
    matricula,
    nome,
    jornada_trabalho,
    cargo,
    setor,
    logins_jms,
    gestor_responsavel,
    folga_dominical,
    genero,
    ativo=1
):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO colaboradores (
            matricula,
            nome,
            jornada_trabalho,
            cargo,
            setor,
            filial,
            logins_jms,
            gestor_responsavel,
            folga_dominical,
            genero,
            ativo
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        str(matricula).strip(),
        str(nome).strip(),
        str(jornada_trabalho).strip(),
        str(cargo).strip(),
        str(setor).strip(),
        FILIAL_PADRAO,
        str(logins_jms).strip(),
        str(gestor_responsavel).strip(),
        str(folga_dominical).strip(),
        str(genero).strip(),
        converter_ativo(ativo)
    ))

    conn.commit()
    conn.close()


def atualizar_colaborador(
    colaborador_id,
    matricula,
    nome,
    jornada_trabalho,
    cargo,
    setor,
    logins_jms,
    gestor_responsavel,
    folga_dominical,
    genero,
    ativo
):
    conn = conectar()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE colaboradores
        SET
            matricula = %s,
            nome = %s,
            jornada_trabalho = %s,
            cargo = %s,
            setor = %s,
            filial = %s,
            logins_jms = %s,
            gestor_responsavel = %s,
            folga_dominical = %s,
            genero = %s,
            ativo = %s
        WHERE id = %s
    """, (
        str(matricula).strip(),
        str(nome).strip(),
        str(jornada_trabalho).strip(),
        str(cargo).strip(),
        str(setor).strip(),
        FILIAL_PADRAO,
        str(logins_jms).strip(),
        str(gestor_responsavel).strip(),
        str(folga_dominical).strip(),
        str(genero).strip(),
        converter_ativo(ativo),
        int(colaborador_id)
    ))

    conn.commit()
    conn.close()


def listar_colaboradores(ativos=True):
    conn = conectar()

    query = """
        SELECT
            id,
            matricula,
            nome,
            jornada_trabalho,
            cargo,
            setor,
            filial,
            logins_jms,
            gestor_responsavel,
            folga_dominical,
            genero,
            ativo
        FROM colaboradores
    """

    if ativos:
        query += " WHERE ativo = 1"

    query += " ORDER BY nome"

    df = pd.read_sql_query(query, conn)
    conn.close()

    if not df.empty:
        df["filial"] = FILIAL_PADRAO

    return df


def salvar_edicao_colaboradores(df):
    conn = conectar()
    cursor = conn.cursor()

    df = df.fillna("")

    atualizados = 0
    inseridos = 0
    ignorados = 0

    for _, row in df.iterrows():
        colaborador_id = row.get("id", "")
        nome = str(row.get("nome", "")).strip()

        if not nome:
            ignorados += 1
            continue

        matricula = str(row.get("matricula", "")).strip()
        jornada_trabalho = str(row.get("jornada_trabalho", "")).strip()
        cargo = str(row.get("cargo", "")).strip()
        setor = str(row.get("setor", "")).strip()
        logins_jms = str(row.get("logins_jms", "")).strip()
        gestor_responsavel = str(row.get("gestor_responsavel", "")).strip()
        folga_dominical = str(row.get("folga_dominical", "")).strip()
        genero = str(row.get("genero", "")).strip()
        ativo = converter_ativo(row.get("ativo", 1))

        if pd.notna(colaborador_id) and str(colaborador_id).strip() != "":
            try:
                colaborador_id = int(float(colaborador_id))

                cursor.execute("""
                    UPDATE colaboradores
                    SET
                        matricula = %s,
                        nome = %s,
                        jornada_trabalho = %s,
                        cargo = %s,
                        setor = %s,
                        filial = %s,
                        logins_jms = %s,
                        gestor_responsavel = %s,
                        folga_dominical = %s,
                        genero = %s,
                        ativo = %s
                    WHERE id = %s
                """, (
                    matricula,
                    nome,
                    jornada_trabalho,
                    cargo,
                    setor,
                    FILIAL_PADRAO,
                    logins_jms,
                    gestor_responsavel,
                    folga_dominical,
                    genero,
                    ativo,
                    colaborador_id
                ))

                atualizados += 1

            except Exception:
                ignorados += 1

        else:
            cursor.execute("""
                INSERT INTO colaboradores (
                    matricula,
                    nome,
                    jornada_trabalho,
                    cargo,
                    setor,
                    filial,
                    logins_jms,
                    gestor_responsavel,
                    folga_dominical,
                    genero,
                    ativo
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                matricula,
                nome,
                jornada_trabalho,
                cargo,
                setor,
                FILIAL_PADRAO,
                logins_jms,
                gestor_responsavel,
                folga_dominical,
                genero,
                ativo
            ))

            inseridos += 1

    conn.commit()
    conn.close()

    return {
        "atualizados": atualizados,
        "inseridos": inseridos,
        "ignorados": ignorados
    }


# =========================
# IMPORTAÇÃO
# =========================

def preparar_dataframe_importacao(df_original):
    df = df_original.copy()

    alias = {
        "matricula": "matricula",
        "matricula do colaborador": "matricula",
        "matrícula": "matricula",
        "id": "matricula",
        "codigo": "matricula",
        "código": "matricula",
        "cod": "matricula",
        "registro": "matricula",

        "nome": "nome",
        "nome do colaborador": "nome",
        "nome colaborador": "nome",
        "colaborador": "nome",
        "funcionario": "nome",
        "funcionário": "nome",
        "empregado": "nome",

        "jornada": "jornada_trabalho",
        "jornada trabalho": "jornada_trabalho",
        "jornada de trabalho": "jornada_trabalho",
        "escala": "jornada_trabalho",
        "horario trabalho": "jornada_trabalho",
        "horário trabalho": "jornada_trabalho",

        "cargo": "cargo",
        "funcao": "cargo",
        "função": "cargo",
        "funcao cargo": "cargo",
        "função cargo": "cargo",
        "ocupacao": "cargo",
        "ocupação": "cargo",

        "setor": "setor",
        "area": "setor",
        "área": "setor",
        "departamento": "setor",

        "logins jms": "logins_jms",
        "login jms": "logins_jms",
        "jms": "logins_jms",
        "logins": "logins_jms",
        "login": "logins_jms",
        "logins - jms": "logins_jms",

        "gestor": "gestor_responsavel",
        "responsavel": "gestor_responsavel",
        "responsável": "gestor_responsavel",
        "nome do responsavel": "gestor_responsavel",
        "nome do responsável": "gestor_responsavel",
        "gestor responsavel": "gestor_responsavel",
        "gestor responsável": "gestor_responsavel",
        "supervisor": "gestor_responsavel",
        "lider": "gestor_responsavel",
        "líder": "gestor_responsavel",

        "folga dominical": "folga_dominical",
        "folga_dominical": "folga_dominical",
        "domingo": "folga_dominical",
        "escala domingo": "folga_dominical",
        "escala dominical": "folga_dominical",

        "genero": "genero",
        "gênero": "genero",
        "sexo": "genero",

        "ativo": "ativo",
        "situacao": "ativo",
        "situação": "ativo",
        "status cadastro": "ativo"
    }

    novas_colunas = []

    for coluna in df.columns:
        coluna_normalizada = normalizar_texto(coluna)
        novas_colunas.append(alias.get(coluna_normalizada, coluna_normalizada))

    df.columns = novas_colunas
    df = df.loc[:, ~df.columns.duplicated()]

    for coluna in COLUNAS_IMPORTACAO:
        if coluna not in df.columns:
            df[coluna] = 1 if coluna == "ativo" else ""

    df = df[COLUNAS_IMPORTACAO]
    df = df.dropna(how="all")
    df = df.fillna("")

    if "ativo" in df.columns:
        df["ativo"] = df["ativo"].apply(converter_ativo).astype(bool)

    return df


def analisar_duplicidades_importacao(df_importado):
    df = df_importado.copy()
    df = df.fillna("")

    base = listar_colaboradores(ativos=False)

    mapa_matricula = {}
    mapa_nome = {}

    if not base.empty:
        for _, row in base.iterrows():
            matricula = normalizar_texto(row.get("matricula", ""))
            nome = normalizar_texto(row.get("nome", ""))

            if matricula:
                mapa_matricula[matricula] = row["id"]

            if nome:
                mapa_nome[nome] = row["id"]

    matricula_norm = df["matricula"].astype(str).apply(normalizar_texto)
    nome_norm = df["nome"].astype(str).apply(normalizar_texto)

    duplicado_matricula_arquivo = matricula_norm.ne("") & matricula_norm.duplicated(keep=False)
    duplicado_nome_arquivo = nome_norm.ne("") & nome_norm.duplicated(keep=False)

    situacoes = []
    ids_existentes = []

    for idx, row in df.iterrows():
        matricula = normalizar_texto(row.get("matricula", ""))
        nome = normalizar_texto(row.get("nome", ""))

        id_existente = ""
        situacao = "Novo"

        if matricula and matricula in mapa_matricula:
            id_existente = mapa_matricula[matricula]
            situacao = "Atualização por matrícula"
        elif nome and nome in mapa_nome:
            id_existente = mapa_nome[nome]
            situacao = "Atualização por nome"

        if duplicado_matricula_arquivo.iloc[idx] or duplicado_nome_arquivo.iloc[idx]:
            situacao = f"{situacao} | Duplicado no arquivo"

        situacoes.append(situacao)
        ids_existentes.append(id_existente)

    df.insert(0, "situacao_importacao", situacoes)
    df.insert(1, "id_existente", ids_existentes)

    return df


def importar_colaboradores(df, atualizar_existentes=True, atualizacao_incremental=True):
    conn = conectar()
    cursor = conn.cursor()

    df = df.fillna("")

    inseridos = 0
    atualizados = 0
    ignorados = 0
    duplicados_arquivo = 0

    if "situacao_importacao" in df.columns:
        duplicados_arquivo = df["situacao_importacao"].astype(str).str.contains(
            "Duplicado no arquivo",
            case=False,
            na=False
        ).sum()

    for _, row in df.iterrows():
        matricula = str(row.get("matricula", "")).strip()
        nome = str(row.get("nome", "")).strip()
        jornada_trabalho = str(row.get("jornada_trabalho", "")).strip()
        cargo = str(row.get("cargo", "")).strip()
        setor = str(row.get("setor", "")).strip()
        logins_jms = str(row.get("logins_jms", "")).strip()
        gestor_responsavel = str(row.get("gestor_responsavel", "")).strip()
        folga_dominical = str(row.get("folga_dominical", "")).strip()
        genero = str(row.get("genero", "")).strip()
        ativo = converter_ativo(row.get("ativo", 1))

        if not nome:
            ignorados += 1
            continue

        existente = None

        if atualizar_existentes:
            id_existente = str(row.get("id_existente", "")).strip()

            if id_existente:
                cursor.execute("""
                    SELECT
                        id,
                        matricula,
                        nome,
                        jornada_trabalho,
                        cargo,
                        setor,
                        logins_jms,
                        gestor_responsavel,
                        folga_dominical,
                        genero,
                        ativo
                    FROM colaboradores
                    WHERE id = %s
                    LIMIT 1
                """, (id_existente,))
                existente = cursor.fetchone()

            if existente is None and matricula:
                cursor.execute("""
                    SELECT
                        id,
                        matricula,
                        nome,
                        jornada_trabalho,
                        cargo,
                        setor,
                        logins_jms,
                        gestor_responsavel,
                        folga_dominical,
                        genero,
                        ativo
                    FROM colaboradores
                    WHERE LOWER(TRIM(matricula)) = LOWER(TRIM(%s))
                    LIMIT 1
                """, (matricula,))
                existente = cursor.fetchone()

            if existente is None:
                cursor.execute("""
                    SELECT
                        id,
                        matricula,
                        nome,
                        jornada_trabalho,
                        cargo,
                        setor,
                        logins_jms,
                        gestor_responsavel,
                        folga_dominical,
                        genero,
                        ativo
                    FROM colaboradores
                    WHERE LOWER(TRIM(nome)) = LOWER(TRIM(%s))
                    LIMIT 1
                """, (nome,))
                existente = cursor.fetchone()

        if existente:
            colaborador_id = existente[0]

            if atualizacao_incremental:
                matricula_final = matricula if matricula else existente[1]
                nome_final = nome if nome else existente[2]
                jornada_final = jornada_trabalho if jornada_trabalho else existente[3]
                cargo_final = cargo if cargo else existente[4]
                setor_final = setor if setor else existente[5]
                logins_final = logins_jms if logins_jms else existente[6]
                gestor_final = gestor_responsavel if gestor_responsavel else existente[7]
                folga_final = folga_dominical if folga_dominical else existente[8]
                genero_final = genero if genero else existente[9]
                ativo_final = ativo
            else:
                matricula_final = matricula
                nome_final = nome
                jornada_final = jornada_trabalho
                cargo_final = cargo
                setor_final = setor
                logins_final = logins_jms
                gestor_final = gestor_responsavel
                folga_final = folga_dominical
                genero_final = genero
                ativo_final = ativo

            cursor.execute("""
                UPDATE colaboradores
                SET
                    matricula = %s,
                    nome = %s,
                    jornada_trabalho = %s,
                    cargo = %s,
                    setor = %s,
                    filial = %s,
                    logins_jms = %s,
                    gestor_responsavel = %s,
                    folga_dominical = %s,
                    genero = %s,
                    ativo = %s
                WHERE id = %s
            """, (
                matricula_final,
                nome_final,
                jornada_final,
                cargo_final,
                setor_final,
                FILIAL_PADRAO,
                logins_final,
                gestor_final,
                folga_final,
                genero_final,
                ativo_final,
                colaborador_id
            ))

            atualizados += 1

        else:
            cursor.execute("""
                INSERT INTO colaboradores (
                    matricula,
                    nome,
                    jornada_trabalho,
                    cargo,
                    setor,
                    filial,
                    logins_jms,
                    gestor_responsavel,
                    folga_dominical,
                    genero,
                    ativo
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                matricula,
                nome,
                jornada_trabalho,
                cargo,
                setor,
                FILIAL_PADRAO,
                logins_jms,
                gestor_responsavel,
                folga_dominical,
                genero,
                ativo
            ))

            inseridos += 1

    conn.commit()
    conn.close()

    return {
        "inseridos": inseridos,
        "atualizados": atualizados,
        "ignorados": ignorados,
        "duplicados_arquivo": int(duplicados_arquivo)
    }


def gerar_modelo_importacao():
    df_modelo = pd.DataFrame([
        {
            "matricula": "1000",
            "nome": "Nome do Colaborador",
            "jornada_trabalho": "09:00 às 17:20 SEGUNDA À SÁBADO",
            "cargo": "Assistente Operacional",
            "setor": "Administrativo",
            "logins_jms": "Sim",
            "gestor_responsavel": "Nome do Responsável",
            "folga_dominical": "Sim",
            "genero": "Feminino",
            "ativo": True
        }
    ])

    arquivo = BytesIO()

    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        df_modelo.to_excel(writer, index=False, sheet_name="Modelo Importação")

    arquivo.seek(0)

    return arquivo


# =========================
# PRESENÇAS
# =========================

def carregar_presencas():
    conn = conectar()

    query = """
        SELECT
            p.id,
            p.colaborador_id,
            p.data,
            c.matricula,
            c.nome,
            c.jornada_trabalho,
            c.cargo,
            c.setor,
            c.filial,
            c.logins_jms,
            c.gestor_responsavel,
            c.folga_dominical,
            c.genero,
            p.status,
            p.observacao,
            p.criado_em,
            p.atualizado_em
        FROM presencas p
        INNER JOIN colaboradores c
            ON p.colaborador_id = c.id
        ORDER BY p.data DESC, c.nome
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    if not df.empty:
        df["filial"] = FILIAL_PADRAO

    return df


def obter_presencas_por_data(data_presenca):
    conn = conectar()

    query = """
        SELECT
            id,
            colaborador_id,
            status,
            observacao
        FROM presencas
        WHERE data = %s
    """

    df = pd.read_sql_query(query, conn, params=(str(data_presenca),))
    conn.close()

    if df.empty:
        return {}

    return df.set_index("colaborador_id").to_dict(orient="index")




def obter_ids_desligados_antes(data_referencia):
    conn = conectar()

    primeiro_dia_mes = date(data_referencia.year, data_referencia.month, 1)

    df = pd.read_sql_query("""
        SELECT DISTINCT colaborador_id
        FROM presencas
        WHERE status = 'Desligado'
          AND data < %s
    """, conn, params=(str(primeiro_dia_mes),))

    conn.close()

    if df.empty:
        return set()

    return set(df["colaborador_id"].tolist())


def calcular_headcount_periodo(colaboradores_base, data_inicio):
    if colaboradores_base.empty:
        return 0

    ids_desligados_antes = obter_ids_desligados_antes(data_inicio)

    base_ativa_periodo = colaboradores_base[
        (colaboradores_base["ativo"].apply(converter_ativo) == 1) &
        (~colaboradores_base["id"].isin(ids_desligados_antes))
    ]

    return len(base_ativa_periodo)

def salvar_presenca(colaborador_id, data_presenca, status, observacao, permitir_alteracao=False):
    colaborador_id = int(colaborador_id)
    conn = conectar()
    cursor = conn.cursor()

    data_formatada = str(data_presenca)

    cursor.execute("""
        SELECT id
        FROM presencas
        WHERE colaborador_id = %s
          AND data = %s
        LIMIT 1
    """, (colaborador_id, data_formatada))

    existente = cursor.fetchone()

    if existente and not permitir_alteracao:
        conn.close()
        return "bloqueado"

    if existente and permitir_alteracao:
        cursor.execute("""
            UPDATE presencas
            SET
                status = %s,
                observacao = %s,
                atualizado_em = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (
            status,
            observacao,
            existente[0]
        ))

        conn.commit()
        conn.close()
        return "atualizado"

    cursor.execute("""
        INSERT INTO presencas (
            colaborador_id,
            data,
            status,
            observacao
        )
        VALUES (%s, %s, %s, %s)
    """, (
        colaborador_id,
        data_formatada,
        status,
        observacao
    ))

    conn.commit()
    conn.close()
    return "inserido"


# =========================
# INDICADORES E HISTÓRICO
# =========================

def calcular_indicadores(df_periodo, colaboradores_base):
    if df_periodo.empty:
        total_registros = 0
        ausencias = 0
        presentes = 0
        faltas = 0
        desligados = 0
    else:
        total_registros = len(df_periodo)
        ausencias = len(df_periodo[df_periodo["status"].isin(STATUS_ABSENTEISMO)])
        presentes = len(df_periodo[df_periodo["status"] == "Presente"])
        faltas = len(df_periodo[df_periodo["status"] == "Falta"])
        desligados = df_periodo[df_periodo["status"] == "Desligado"]["colaborador_id"].nunique()

    headcount = len(colaboradores_base[colaboradores_base["ativo"] == 1]) if not colaboradores_base.empty else 0

    abs_pct = (ausencias / total_registros * 100) if total_registros > 0 else 0
    turnover_pct = (desligados / headcount * 100) if headcount > 0 else 0

    return {
        "total_registros": total_registros,
        "presentes": presentes,
        "faltas": faltas,
        "ausencias": ausencias,
        "desligados": desligados,
        "headcount": headcount,
        "abs_pct": abs_pct,
        "turnover_pct": turnover_pct
    }


def montar_tabela_historico(df):
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], errors="coerce")
    df = df.dropna(subset=["data_dt"])
    df["data_coluna"] = df["data_dt"].dt.strftime("%d/%m/%Y")

    colunas_base = [
        "nome",
        "matricula",
        "jornada_trabalho",
        "cargo",
        "setor",
        "logins_jms"
    ]

    tabela = df.pivot_table(
        index=colunas_base,
        columns="data_coluna",
        values="status",
        aggfunc="last"
    ).reset_index()

    datas_ordenadas = sorted(
        [col for col in tabela.columns if col not in colunas_base],
        key=lambda x: pd.to_datetime(x, format="%d/%m/%Y")
    )

    tabela = tabela[colunas_base + datas_ordenadas]

    tabela = tabela.rename(columns={
        "nome": "Nome do Colaborador",
        "matricula": t("Matrícula"),
        "jornada_trabalho": t("Jornada de Trabalho"),
        "cargo": t("Cargo"),
        "setor": t("Setor"),
        "logins_jms": "Logins - JMS"
    })

    return tabela


def montar_ofensores_abs(df_periodo):
    if df_periodo.empty:
        return pd.DataFrame()

    df_abs = df_periodo[df_periodo["status"].isin(STATUS_ABSENTEISMO)].copy()

    if df_abs.empty:
        return pd.DataFrame()

    registros_totais = (
        df_periodo.groupby(["colaborador_id", "nome"])
        .size()
        .reset_index(name="registros_periodo")
    )

    ocorrencias_abs = (
        df_abs.groupby(["colaborador_id", "nome"])
        .size()
        .reset_index(name="ocorrencias_abs")
    )

    ofensores = ocorrencias_abs.merge(
        registros_totais,
        on=["colaborador_id", "nome"],
        how="left"
    )

    ofensores["taxa_abs_colaborador"] = (
        ofensores["ocorrencias_abs"] / ofensores["registros_periodo"] * 100
    )

    ofensores = ofensores.sort_values(
        ["ocorrencias_abs", "taxa_abs_colaborador"],
        ascending=[False, False]
    )

    return ofensores


def montar_headcount_por_responsavel(colaboradores):
    if colaboradores.empty:
        return pd.DataFrame()

    df = colaboradores.copy()

    if "ativo" in df.columns:
        df = df[df["ativo"].apply(converter_ativo) == 1]

    df["gestor_responsavel"] = df["gestor_responsavel"].fillna("").astype(str).str.strip()
    df.loc[df["gestor_responsavel"] == "", "gestor_responsavel"] = "Sem responsável"

    resumo = (
        df.groupby("gestor_responsavel")
        .size()
        .reset_index(name="quantidade_pessoas")
        .sort_values("quantidade_pessoas", ascending=False)
    )

    return resumo


def estilo_status(valor):
    if valor == "Presente":
        return "background-color: #C6EFCE; color: #006100"
    if valor == "Falta":
        return "background-color: #FFC7CE; color: #9C0006"
    if valor == "Atestado":
        return "background-color: #DDEBF7; color: #1F4E78"
    if valor in ["Folga", t("Folga Dominical"), "DSR", "Férias", "Feriado"]:
        return "background-color: #FFF2CC; color: #7F6000"
    if valor in ["Afastamento", "Desligado"]:
        return "background-color: #E7E6E6; color: #404040"
    if valor == "Integração":
        return "background-color: #E2F0D9; color: #375623"
    return ""


# =========================
# EXPORTAÇÃO EXCEL
# =========================

def gerar_excel_historico(df_historico):
    arquivo = BytesIO()

    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        df_historico.to_excel(writer, index=False, sheet_name="Histórico")

    arquivo.seek(0)
    return arquivo


def gerar_excel_modelo(df, data_ref, responsavel):
    caminho_template = Path(TEMPLATE_EXCEL)

    if caminho_template.exists():
        wb = load_workbook(caminho_template)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Presença RH"

    ws["A1"] = "LISTA DE PRESENÇA RH"
    ws["A2"] = f"FILIAL: {FILIAL_PADRAO}"
    ws["A3"] = f"DATA: {data_ref.strftime('%d/%m/%Y')}"
    ws["A4"] = f"RESPONSÁVEL: {responsavel}"

    ws["A6"] = "Nº"
    ws["B6"] = "Nome do Colaborador"
    ws["C6"] = t("Matrícula")
    ws["D6"] = t("Jornada de Trabalho")
    ws["E6"] = t("Cargo")
    ws["F6"] = t("Setor")
    ws["G6"] = "Logins - JMS"
    ws["H6"] = t("Folga Dominical")
    ws["I6"] = t("Gênero")
    ws["J6"] = t("Status")
    ws["K6"] = t("Observação")

    linha_inicial = 7
    df = df.reset_index(drop=True)

    for idx, row in df.iterrows():
        linha = linha_inicial + idx

        ws[f"A{linha}"] = idx + 1
        ws[f"B{linha}"] = row.get("nome", "")
        ws[f"C{linha}"] = row.get("matricula", "")
        ws[f"D{linha}"] = row.get("jornada_trabalho", "")
        ws[f"E{linha}"] = row.get("cargo", "")
        ws[f"F{linha}"] = row.get("setor", "")
        ws[f"G{linha}"] = row.get("logins_jms", "")
        ws[f"H{linha}"] = row.get("folga_dominical", "")
        ws[f"I{linha}"] = row.get("genero", "")
        ws[f"J{linha}"] = row.get("status", "")
        ws[f"K{linha}"] = row.get("observacao", "")

    arquivo = BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)

    return arquivo


# =========================
# LOGIN
# =========================

def login():
    st.title(t("Lista de Presença RH"))
    seletor_idioma("idioma_login")

    usuario = st.text_input(t("Usuário"))
    senha = st.text_input(t("Senha"), type="password")

    if st.button(t("Entrar")):
        dados_usuario = autenticar_usuario(usuario, senha)

        if dados_usuario:
            st.session_state["logado"] = True
            st.session_state["usuario"] = dados_usuario["usuario"]
            st.session_state["perfil"] = dados_usuario["perfil"]
            st.rerun()
        else:
            st.error(t("Usuário ou senha inválidos."))



# =========================
# ESCALA DE FOLGA DOMINICAL
# =========================

def valor_positivo_folga_dominical(valor):
    texto = normalizar_texto(valor)
    return texto in ["sim", "s", "1", "true", "verdadeiro", "folga", "folga dominical"]


def normalizar_genero_operacional(valor):
    texto = normalizar_texto(valor)
    if texto in ["f", "fem", "feminino", "mulher"]:
        return "Feminino"
    if texto in ["m", "masc", "masculino", "homem"]:
        return "Masculino"
    if texto in ["outro", "outros", "nao informado", "não informado"]:
        return "Outro"
    return str(valor).strip() if str(valor).strip() else "Não informado"


def domingos_do_mes(ano, mes):
    primeiro_dia = date(int(ano), int(mes), 1)
    if mes == 12:
        proximo_mes = date(int(ano) + 1, 1, 1)
    else:
        proximo_mes = date(int(ano), int(mes) + 1, 1)

    dias = []
    dia = primeiro_dia
    while dia < proximo_mes:
        if dia.weekday() == 6:
            dias.append(dia)
        dia += timedelta(days=1)
    return dias


def obter_historico_folga_dominical(data_limite):
    conn = conectar()
    df = pd.read_sql_query("""
        SELECT colaborador_id, data
        FROM presencas
        WHERE status = 'Folga Dominical'
          AND data < %s
        ORDER BY colaborador_id, data DESC
    """, conn, params=(str(data_limite),))
    conn.close()

    if df.empty:
        return {}

    df["data"] = pd.to_datetime(df["data"], errors="coerce").dt.date
    df = df.dropna(subset=["data"])
    return df.groupby("colaborador_id")["data"].max().to_dict()


def obter_escala_folga_dominical(data_inicio, data_fim, responsavel="Todos"):
    conn = conectar()
    query = """
        SELECT
            e.id,
            e.colaborador_id,
            e.data_domingo,
            c.matricula,
            c.nome,
            c.cargo,
            c.setor,
            c.gestor_responsavel,
            c.folga_dominical,
            c.genero,
            e.status,
            e.motivo_regra,
            e.criado_em,
            e.atualizado_em
        FROM escalas_folga_dominical e
        INNER JOIN colaboradores c ON c.id = e.colaborador_id
        WHERE e.data_domingo BETWEEN %s AND %s
    """
    params = [str(data_inicio), str(data_fim)]

    if responsavel != "Todos":
        query += " AND c.gestor_responsavel = %s"
        params.append(responsavel)

    query += " ORDER BY e.data_domingo, c.gestor_responsavel, c.nome"

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def gerar_escala_folga_dominical(colaboradores_base, ano, mes, responsavel="Todos"):
    domingos = domingos_do_mes(ano, mes)

    if not domingos or colaboradores_base.empty:
        return pd.DataFrame()

    primeiro_dia_mes = date(int(ano), int(mes), 1)
    ids_desligados_antes = obter_ids_desligados_antes(primeiro_dia_mes)

    df = colaboradores_base.copy()
    df = df[~df["id"].isin(ids_desligados_antes)].copy()

    if responsavel != "Todos":
        df = df[df["gestor_responsavel"] == responsavel].copy()

    if df.empty:
        return pd.DataFrame()

    df["folga_dominical_flag"] = df["folga_dominical"].apply(valor_positivo_folga_dominical)
    df = df[df["folga_dominical_flag"]].copy()

    if df.empty:
        return pd.DataFrame()

    historico_ultima_folga = obter_historico_folga_dominical(primeiro_dia_mes)
    registros = []

    for _, row in df.iterrows():
        colaborador_id = row["id"]
        genero = normalizar_genero_operacional(row.get("genero", ""))
        ultima_folga = historico_ultima_folga.get(colaborador_id)

        if genero == "Feminino":
            domingos_colaborador = []
            for domingo in domingos:
                if ultima_folga is None or (domingo - ultima_folga).days >= 14:
                    domingos_colaborador.append(domingo)
                    ultima_folga = domingo
            motivo = "Regra operacional: folga dominical em ciclo quinzenal para gênero feminino."
        else:
            domingos_colaborador = [domingos[0]]
            motivo = "Regra operacional: uma folga dominical mensal, sem limite por responsável."

        for domingo in domingos_colaborador:
            registros.append({
                "data_domingo": domingo,
                "matricula": row.get("matricula", ""),
                "nome": row.get("nome", ""),
                "cargo": row.get("cargo", ""),
                "setor": row.get("setor", ""),
                "responsavel": row.get("gestor_responsavel", ""),
                "folga_dominical": row.get("folga_dominical", ""),
                "genero": genero,
                "status": t("Folga Dominical"),
                "motivo_regra": motivo,
                "colaborador_id": colaborador_id
            })

    escala = pd.DataFrame(registros)

    if escala.empty:
        return escala

    return escala.sort_values(["data_domingo", "responsavel", "nome"]).reset_index(drop=True)


def salvar_escala_folga_dominical(df_escala):
    if df_escala.empty:
        return {"gerados": 0, "presencas_atualizadas": 0}

    conn = conectar()
    cursor = conn.cursor()
    gerados = 0
    presencas_atualizadas = 0

    for _, row in df_escala.iterrows():
        colaborador_id = int(row["colaborador_id"])
        data_domingo = str(row["data_domingo"])
        motivo = str(row.get("motivo_regra", ""))
        observacao = "Gerado pela escala de folga dominical."

        cursor.execute("""
            INSERT INTO escalas_folga_dominical (
                colaborador_id,
                data_domingo,
                status,
                motivo_regra
            )
            VALUES (%s, %s, 'Folga Dominical', %s)
            ON CONFLICT (colaborador_id, data_domingo)
            DO UPDATE SET
                status = 'Folga Dominical',
                motivo_regra = EXCLUDED.motivo_regra,
                atualizado_em = CURRENT_TIMESTAMP
        """, (colaborador_id, data_domingo, motivo))
        gerados += 1

        cursor.execute("""
            INSERT INTO presencas (
                colaborador_id,
                data,
                status,
                observacao
            )
            VALUES (%s, %s, 'Folga Dominical', %s)
            ON CONFLICT (colaborador_id, data)
            DO UPDATE SET
                status = 'Folga Dominical',
                observacao = EXCLUDED.observacao,
                atualizado_em = CURRENT_TIMESTAMP
        """, (colaborador_id, data_domingo, observacao))
        presencas_atualizadas += 1

    conn.commit()
    conn.close()

    return {"gerados": gerados, "presencas_atualizadas": presencas_atualizadas}


def gerar_excel_escala_folga_dominical(df_escala):
    arquivo = BytesIO()
    df_export = df_escala.copy()

    if not df_export.empty:
        df_export["data_domingo"] = pd.to_datetime(df_export["data_domingo"]).dt.strftime("%d/%m/%Y")
        df_export = df_export.rename(columns={
            "data_domingo": "Data Domingo",
            "matricula": t("Matrícula"),
            "nome": t("Nome"),
            "cargo": t("Cargo"),
            "setor": t("Setor"),
            "responsavel": t("Responsável"),
            "folga_dominical": t("Folga Dominical"),
            "genero": t("Gênero"),
            "status": t("Status"),
            "motivo_regra": "Regra Aplicada"
        })
        colunas = [
            "Data Domingo", t("Matrícula"), t("Nome"), t("Cargo"), t("Setor"),
            t("Responsável"), t("Folga Dominical"), t("Gênero"), t("Status"), "Regra Aplicada"
        ]
        df_export = df_export[[col for col in colunas if col in df_export.columns]]

    with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Escala Folga Dominical")

    arquivo.seek(0)
    return arquivo


def gerar_pdf_escala_folga_dominical(df_escala, titulo="Escala de Folga Dominical"):
    arquivo = BytesIO()
    doc = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18
    )

    styles = getSampleStyleSheet()
    elementos = [Paragraph(titulo, styles["Title"]), Spacer(1, 10)]

    if df_escala.empty:
        elementos.append(Paragraph("Nenhum registro encontrado para os filtros selecionados.", styles["Normal"]))
    else:
        df_pdf = df_escala.copy()
        df_pdf["data_domingo"] = pd.to_datetime(df_pdf["data_domingo"]).dt.strftime("%d/%m/%Y")
        colunas = ["data_domingo", "matricula", "nome", "cargo", "responsavel", "genero", "status"]
        headers = [t("Data"), t("Matrícula"), t("Nome"), t("Cargo"), t("Responsável"), t("Gênero"), t("Status")]
        dados = [headers]

        for _, row in df_pdf[colunas].iterrows():
            dados.append([str(row.get(col, "")) for col in colunas])

        tabela = Table(dados, repeatRows=1, colWidths=[58, 62, 165, 130, 100, 70, 95])
        tabela.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ]))
        elementos.append(tabela)

    doc.build(elementos)
    arquivo.seek(0)
    return arquivo


def pagina_escala_folga_dominical(modo="Gestor"):
    st.subheader(t("Escala de Folga Dominical"))
    st.caption("A geração usa a base de colaboradores, as colunas folga_dominical e gênero, e o histórico de Folga Dominical já salvo.")

    colaboradores = listar_colaboradores(ativos=True)

    if colaboradores.empty:
        st.warning("Nenhum colaborador cadastrado.")
        return

    hoje = date.today()
    col1, col2, col3 = st.columns(3)

    with col1:
        ano = st.number_input(t("Ano"), min_value=2020, max_value=2100, value=hoje.year, step=1, key=f"escala_ano_{modo}")

    with col2:
        mes = st.selectbox(
            t("Mês"),
            list(range(1, 13)),
            index=hoje.month - 1,
            format_func=lambda x: f"{x:02d}",
            key=f"escala_mes_{modo}"
        )

    with col3:
        responsavel = st.selectbox(
            t("Responsável"),
            opcoes_unicas(colaboradores, "gestor_responsavel", "Todos"),
            key=f"escala_responsavel_{modo}"
        )

    primeiro_dia = date(int(ano), int(mes), 1)
    if int(mes) == 12:
        ultimo_dia = date(int(ano), 12, 31)
    else:
        ultimo_dia = date(int(ano), int(mes) + 1, 1) - timedelta(days=1)

    if modo == "Gestor":
        escala_previa = gerar_escala_folga_dominical(colaboradores, int(ano), int(mes), responsavel)

        if escala_previa.empty:
            st.warning("Nenhuma escala gerada. Verifique se a coluna folga_dominical está como Sim para os colaboradores elegíveis.")
            return

        st.metric("Registros previstos", len(escala_previa))

        st.dataframe(
            escala_previa.drop(columns=["colaborador_id"], errors="ignore"),
            use_container_width=True,
            hide_index=True
        )

        colb1, colb2, colb3 = st.columns(3)

        with colb1:
            if st.button("Gravar escala no histórico", key=f"salvar_escala_{modo}"):
                resultado = salvar_escala_folga_dominical(escala_previa)
                st.success(
                    f"Escala gravada. Registros da escala: {resultado['gerados']} | "
                    f"Presenças atualizadas: {resultado['presencas_atualizadas']}"
                )
                st.rerun()

        with colb2:
            excel = gerar_excel_escala_folga_dominical(escala_previa)
            st.download_button(
                "Baixar Excel",
                data=excel,
                file_name=f"escala_folga_dominical_{ano}_{int(mes):02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_escala_{modo}"
            )

        with colb3:
            pdf = gerar_pdf_escala_folga_dominical(
                escala_previa,
                titulo=f"Escala de Folga Dominical - {int(mes):02d}/{ano}"
            )
            st.download_button(
                "Baixar PDF paisagem",
                data=pdf,
                file_name=f"escala_folga_dominical_{ano}_{int(mes):02d}.pdf",
                mime="application/pdf",
                key=f"download_pdf_escala_{modo}"
            )

    else:
        escala_salva = obter_escala_folga_dominical(primeiro_dia, ultimo_dia, responsavel)

        if escala_salva.empty:
            st.warning("Nenhuma escala salva para o período selecionado.")
            return

        st.metric("Registros salvos", len(escala_salva))
        st.dataframe(escala_salva, use_container_width=True, hide_index=True)

        excel = gerar_excel_escala_folga_dominical(escala_salva.rename(columns={"data_domingo": "data_domingo", "gestor_responsavel": "responsavel"}))
        pdf = gerar_pdf_escala_folga_dominical(
            escala_salva.rename(columns={"gestor_responsavel": "responsavel"}),
            titulo=f"Escala de Folga Dominical - {int(mes):02d}/{ano}"
        )

        colb1, colb2 = st.columns(2)
        with colb1:
            st.download_button(
                "Baixar Excel",
                data=excel,
                file_name=f"escala_folga_dominical_{ano}_{int(mes):02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_escala_{modo}"
            )
        with colb2:
            st.download_button(
                "Baixar PDF paisagem",
                data=pdf,
                file_name=f"escala_folga_dominical_{ano}_{int(mes):02d}.pdf",
                mime="application/pdf",
                key=f"download_pdf_escala_{modo}"
            )

# =========================
# OPERAÇÃO
# =========================

def pagina_operacao_lancamento():
    st.subheader(t("Lançamento de Presença"))
    st.caption(f"Filial padrão: {FILIAL_PADRAO}")

    colaboradores = listar_colaboradores(ativos=True)

    if colaboradores.empty:
        st.warning("Nenhum colaborador cadastrado.")
        return

    col1, col2 = st.columns([1, 2])

    with col1:
        data_presenca = st.date_input(t("Data"), value=date.today())

    with col2:
        responsavel_selecionado = st.selectbox(
            t("Responsável"),
            opcoes_unicas(colaboradores, "gestor_responsavel", "Todos")
        )

    ids_desligados_antes = obter_ids_desligados_antes(data_presenca)

    colaboradores = colaboradores[
        ~colaboradores["id"].isin(ids_desligados_antes)
    ].copy()

    if colaboradores.empty:
        st.warning("Nenhum colaborador ativo para lançamento nesta competência.")
        return

    colaboradores_filtrados = colaboradores.copy()

    if responsavel_selecionado != "Todos":
        colaboradores_filtrados = colaboradores_filtrados[
            colaboradores_filtrados["gestor_responsavel"] == responsavel_selecionado
        ]

    if colaboradores_filtrados.empty:
        st.warning("Nenhum colaborador encontrado para esse responsável.")
        return

    presencas_dia = obter_presencas_por_data(data_presenca)
    ids_filtrados = set(colaboradores_filtrados["id"].tolist())
    ids_ja_lancados = ids_filtrados.intersection(set(presencas_dia.keys()))

    total_filtrados = len(colaboradores_filtrados)
    total_lancados = len(ids_ja_lancados)

    colm1, colm2, colm3 = st.columns(3)
    colm1.metric("Colaboradores na lista", total_filtrados)
    colm2.metric("Já lançados na data", total_lancados)
    colm3.metric("Pendentes", total_filtrados - total_lancados)

    chave_alteracao = f"alteracao_{data_presenca}_{responsavel_selecionado}"

    if total_lancados > 0:
        st.warning(
            "Existem registros já salvos para essa data. "
            "Eles ficam bloqueados para evitar duplicidade. Para corrigir, clique em Alterar."
        )

        if st.button("Alterar lançamentos desta data"):
            st.session_state[chave_alteracao] = True
            st.rerun()

    modo_alteracao = st.session_state.get(chave_alteracao, False)

    if modo_alteracao:
        st.info("Modo de alteração ativo. Ao salvar, os registros existentes serão atualizados.")

    st.divider()

    with st.form("form_lancamento_presenca"):
        col_h1, col_h2, col_h3, col_h4, col_h5, col_h6 = st.columns(
            [1.1, 3, 2.2, 2.2, 2, 3]
        )

        col_h1.markdown("**Matrícula**")
        col_h2.markdown("**Nome**")
        col_h3.markdown("**Cargo**")
        col_h4.markdown("**Responsável**")
        col_h5.markdown("**Status**")
        col_h6.markdown("**Observação**")

        registros = []

        for _, row in colaboradores_filtrados.iterrows():
            existente = presencas_dia.get(row["id"])
            ja_lancado = existente is not None
            bloqueado = ja_lancado and not modo_alteracao

            status_atual = existente["status"] if existente else None

            index_status = (
                STATUS_PRESENCA.index(status_atual)
                if status_atual in STATUS_PRESENCA
                else None
            )

            observacao_atual = existente["observacao"] if existente else ""
            observacao_atual = "" if pd.isna(observacao_atual) else str(observacao_atual)

            col1, col2, col3, col4, col5, col6 = st.columns(
                [1.1, 3, 2.2, 2.2, 2, 3]
            )

            col1.write(row["matricula"])
            col2.write(row["nome"])
            col3.write(row["cargo"])
            col4.write(row["gestor_responsavel"])

            with col5:
                status = st.selectbox(
                    t("Status"),
                    STATUS_PRESENCA,
                    index=index_status,
                    placeholder="Selecione",
                    key=f"status_{row['id']}_{data_presenca}_{responsavel_selecionado}",
                    label_visibility="collapsed",
                    disabled=bloqueado
                )

            with col6:
                observacao = st.text_input(
                    t("Observação"),
                    value=observacao_atual,
                    key=f"obs_{row['id']}_{data_presenca}_{responsavel_selecionado}",
                    label_visibility="collapsed",
                    disabled=bloqueado
                )

            registros.append({
                "colaborador_id": row["id"],
                "status": status,
                "observacao": observacao,
                "ja_lancado": ja_lancado
            })

        pode_salvar = modo_alteracao or total_lancados < total_filtrados
        rotulo_botao = "Salvar alterações" if modo_alteracao else "Salvar novos lançamentos"

        salvar = st.form_submit_button(rotulo_botao, disabled=not pode_salvar)

        if salvar:
            if any(item["status"] is None for item in registros):
                st.warning("Selecione o status de todos os colaboradores antes de salvar.")
                st.stop()

            inseridos = 0
            atualizados = 0
            bloqueados = 0

            for item in registros:
                resultado = salvar_presenca(
                    item["colaborador_id"],
                    data_presenca,
                    item["status"],
                    item["observacao"],
                    permitir_alteracao=modo_alteracao
                )

                if resultado == "inserido":
                    inseridos += 1
                elif resultado == "atualizado":
                    atualizados += 1
                elif resultado == "bloqueado":
                    bloqueados += 1

            st.success(
                f"Processamento concluído. "
                f"Inseridos: {inseridos} | Atualizados: {atualizados} | Bloqueados: {bloqueados}"
            )

            st.session_state[chave_alteracao] = False
            st.rerun()


def pagina_operacao_historico():
    st.subheader(t("Histórico de Presença"))
    st.caption(f"Filial padrão: {FILIAL_PADRAO}")

    colaboradores = listar_colaboradores(ativos=True)
    df = carregar_presencas()

    if colaboradores.empty:
        st.warning("Nenhum colaborador cadastrado.")
        return

    col1, col2, col3 = st.columns(3)

    with col1:
        data_inicio = st.date_input(
            t("Data inicial"),
            value=date.today().replace(day=1),
            key="hist_data_inicio_op"
        )

    with col2:
        data_fim = st.date_input(
            t("Data final"),
            value=date.today(),
            key="hist_data_fim_op"
        )

    with col3:
        responsavel = st.selectbox(
            t("Responsável"),
            opcoes_unicas(colaboradores, "gestor_responsavel", "Todos"),
            key="hist_responsavel_op"
        )

    colaboradores_base = colaboradores.copy()

    if responsavel != "Todos":
        colaboradores_base = colaboradores_base[
            colaboradores_base["gestor_responsavel"] == responsavel
        ]

    if df.empty:
        df_periodo = pd.DataFrame()
    else:
        df["data_dt"] = pd.to_datetime(df["data"], errors="coerce").dt.date

        df_periodo = df[
            (df["data_dt"] >= data_inicio) &
            (df["data_dt"] <= data_fim)
        ].copy()

        if responsavel != "Todos":
            df_periodo = df_periodo[df_periodo["gestor_responsavel"] == responsavel]

    indicadores = calcular_indicadores(df_periodo, colaboradores_base)
    indicadores["headcount"] = calcular_headcount_periodo(colaboradores_base, data_inicio)
    indicadores["turnover_pct"] = (
        indicadores["desligados"] / indicadores["headcount"] * 100
        if indicadores["headcount"] > 0
        else 0
    )

    colm1, colm2, colm3, colm4 = st.columns(4)
    colm1.metric("Headcount", indicadores["headcount"])
    colm2.metric("ABS", f"{indicadores['abs_pct']:.1f}%")
    colm3.metric("Turnover", f"{indicadores['turnover_pct']:.1f}%")
    colm4.metric("Desligados", indicadores["desligados"])

    colm5, colm6, colm7, colm8 = st.columns(4)
    colm5.metric("Registros", indicadores["total_registros"])
    colm6.metric("Presentes", indicadores["presentes"])
    colm7.metric("Faltas", indicadores["faltas"])
    colm8.metric("Ausências ABS", indicadores["ausencias"])

    st.caption(
        "ABS considera Falta, Atestado e Afastamento. "
        "Turnover considera status Desligado dividido pelo headcount ativo do filtro."
    )

    st.divider()

    if df_periodo.empty:
        st.warning("Nenhum lançamento encontrado para o período selecionado.")
        return

    tabela_historico = montar_tabela_historico(df_periodo)

    st.dataframe(
        tabela_historico.style.map(estilo_status),
        use_container_width=True,
        hide_index=True
    )

    arquivo_historico = gerar_excel_historico(tabela_historico)

    st.download_button(
        "Baixar histórico em Excel",
        data=arquivo_historico,
        file_name=f"historico_presenca_{data_inicio}_a_{data_fim}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def pagina_operacao():
    st.title(t("Operação - Presença RH"))

    aba1, aba2, aba3 = st.tabs([
        t("Lançamento"),
        t("Histórico"),
        t("Escala Folga Dominical")
    ])

    with aba1:
        pagina_operacao_lancamento()

    with aba2:
        pagina_operacao_historico()

    with aba3:
        pagina_escala_folga_dominical(modo="Operação")


# =========================
# GESTOR
# =========================

def pagina_gestor():
    st.title(t("Painel do Gestor - RH"))

    aba1, aba2, aba3, aba4, aba5, aba6, aba7 = st.tabs([
        t("Dashboard"),
        t("Importar Dados"),
        t("Cadastrar / Alterar Pessoas"),
        t("Colaboradores"),
        t("Escala Folga Dominical"),
        t("Exportar Excel"),
        t("Acessos")
    ])

    with aba1:
        st.subheader(t("Dashboard gerencial"))
        st.caption(f"Filial padrão: {FILIAL_PADRAO}")

        colaboradores = listar_colaboradores(ativos=True)
        df = carregar_presencas()

        colf1, colf2, colf3 = st.columns(3)

        with colf1:
            data_inicio = st.date_input(
                t("Data inicial"),
                value=date.today().replace(day=1),
                key="dash_data_inicio"
            )

        with colf2:
            data_fim = st.date_input(
                t("Data final"),
                value=date.today(),
                key="dash_data_fim"
            )

        with colf3:
            responsavel = st.selectbox(
                t("Responsável"),
                opcoes_unicas(colaboradores, "gestor_responsavel", "Todos"),
                key="dash_responsavel"
            )

        colaboradores_base = colaboradores.copy()

        if responsavel != "Todos":
            colaboradores_base = colaboradores_base[
                colaboradores_base["gestor_responsavel"] == responsavel
            ]

        if df.empty:
            df_periodo = pd.DataFrame()
        else:
            df["data_dt"] = pd.to_datetime(df["data"], errors="coerce").dt.date

            df_periodo = df[
                (df["data_dt"] >= data_inicio) &
                (df["data_dt"] <= data_fim)
            ].copy()

            if responsavel != "Todos":
                df_periodo = df_periodo[df_periodo["gestor_responsavel"] == responsavel]

        indicadores = calcular_indicadores(df_periodo, colaboradores_base)

        colm1, colm2, colm3, colm4 = st.columns(4)
        colm1.metric("Headcount", indicadores["headcount"])
        colm2.metric("ABS", f"{indicadores['abs_pct']:.1f}%")
        colm3.metric("Turnover", f"{indicadores['turnover_pct']:.1f}%")
        colm4.metric("Desligados", indicadores["desligados"])

        colm5, colm6, colm7, colm8 = st.columns(4)
        colm5.metric("Registros", indicadores["total_registros"])
        colm6.metric("Presentes", indicadores["presentes"])
        colm7.metric("Faltas", indicadores["faltas"])
        colm8.metric("Ausências ABS", indicadores["ausencias"])

        st.caption(
            "ABS considera Falta, Atestado e Afastamento. "
            "Turnover considera status Desligado dividido pelo headcount ativo do filtro."
        )

        st.divider()

        st.subheader("Quantidade total de pessoas por responsável")

        headcount_responsavel = montar_headcount_por_responsavel(colaboradores)

        if headcount_responsavel.empty:
            st.warning("Nenhum colaborador ativo encontrado para montar o gráfico.")
        else:
            grafico_headcount = alt.Chart(headcount_responsavel).mark_bar().encode(
                x=alt.X(
                    "gestor_responsavel:N",
                    sort=alt.EncodingSortField(
                        field="quantidade_pessoas",
                        order="descending"
                    ),
                    title=t("Responsável"),
                    axis=alt.Axis(labelAngle=-45)
                ),
                y=alt.Y(
                    "quantidade_pessoas:Q",
                    title="Quantidade de pessoas"
                ),
                tooltip=[
                    alt.Tooltip("gestor_responsavel:N", title=t("Responsável")),
                    alt.Tooltip("quantidade_pessoas:Q", title="Quantidade de pessoas")
                ]
            ).properties(
                height=420
            )

            st.altair_chart(grafico_headcount, use_container_width=True)

            st.dataframe(
                headcount_responsavel.rename(columns={
                    "gestor_responsavel": t("Responsável"),
                    "quantidade_pessoas": "Quantidade de pessoas"
                }),
                use_container_width=True,
                hide_index=True
            )

        if df_periodo.empty:
            st.warning("Nenhum lançamento encontrado para o período selecionado.")
        else:
            st.divider()

            resumo_status = (
                df_periodo.groupby("status")
                .size()
                .reset_index(name="quantidade")
                .sort_values("quantidade", ascending=False)
            )

            st.subheader("Volume por Status")

            grafico_status = alt.Chart(resumo_status).mark_bar().encode(
                x=alt.X(
                    "status:N",
                    sort=alt.EncodingSortField(
                        field="quantidade",
                        order="descending"
                    ),
                    title=t("Status"),
                    axis=alt.Axis(labelAngle=-45)
                ),
                y=alt.Y(
                    "quantidade:Q",
                    title="Quantidade"
                ),
                tooltip=[
                    alt.Tooltip("status:N", title=t("Status")),
                    alt.Tooltip("quantidade:Q", title="Quantidade")
                ]
            ).properties(
                height=420
            )

            st.altair_chart(grafico_status, use_container_width=True)

            st.subheader("Colaboradores ofensores em ABS")

            ofensores_abs = montar_ofensores_abs(df_periodo)

            if ofensores_abs.empty:
                st.info("Nenhuma ocorrência de ABS no período selecionado.")
            else:
                col_top_abs, col_total_abs = st.columns([1, 3])

                with col_top_abs:
                    filtro_top_abs = st.selectbox(
                        "Filtro do ranking",
                        ["Top 3", "Top 5", "Top 10", "Todos"],
                        index=2,
                        key="filtro_top_ofensores_abs"
                    )

                mapa_top_abs = {
                    "Top 3": 3,
                    "Top 5": 5,
                    "Top 10": 10
                }

                limite_top_abs = mapa_top_abs.get(filtro_top_abs)

                if limite_top_abs:
                    ofensores_grafico = ofensores_abs.head(limite_top_abs).copy()
                else:
                    ofensores_grafico = ofensores_abs.copy()

                with col_total_abs:
                    st.metric(
                        "Colaboradores exibidos no ranking",
                        len(ofensores_grafico)
                    )

                grafico_abs = alt.Chart(ofensores_grafico).mark_bar().encode(
                    x=alt.X(
                        "nome:N",
                        sort=alt.EncodingSortField(
                            field="ocorrencias_abs",
                            order="descending"
                        ),
                        title="Colaborador",
                        axis=alt.Axis(labelAngle=-45)
                    ),
                    y=alt.Y(
                        "ocorrencias_abs:Q",
                        title="Ocorrências ABS"
                    ),
                    tooltip=[
                        alt.Tooltip("nome:N", title="Colaborador"),
                        alt.Tooltip("ocorrencias_abs:Q", title="Ocorrências ABS"),
                        alt.Tooltip("registros_periodo:Q", title="Registros no período"),
                        alt.Tooltip("taxa_abs_colaborador:Q", title="Taxa ABS individual", format=".1f")
                    ]
                ).properties(
                    height=420
                )

                st.altair_chart(grafico_abs, use_container_width=True)

                st.dataframe(
                    ofensores_grafico.rename(columns={
                        "nome": "Colaborador",
                        "ocorrencias_abs": "Ocorrências ABS",
                        "registros_periodo": "Registros no período",
                        "taxa_abs_colaborador": "Taxa ABS individual %"
                    }),
                    use_container_width=True,
                    hide_index=True
                )

            st.subheader("Resumo por Responsável")

            resumo_responsavel = (
                df_periodo.groupby(["gestor_responsavel", "status"])
                .size()
                .reset_index(name="quantidade")
                .sort_values(["gestor_responsavel", "quantidade"], ascending=[True, False])
            )

            st.dataframe(
                resumo_responsavel,
                use_container_width=True,
                hide_index=True
            )

            st.subheader("Histórico consolidado")

            tabela_historico = montar_tabela_historico(df_periodo)

            st.dataframe(
                tabela_historico.style.map(estilo_status),
                use_container_width=True,
                hide_index=True
            )

    with aba2:
        st.subheader(t("Importar dados de colaboradores"))
        st.caption(f"Filial padrão aplicada automaticamente: {FILIAL_PADRAO}")

        modelo_importacao = gerar_modelo_importacao()

        st.download_button(
            "Baixar modelo de importação",
            data=modelo_importacao,
            file_name="modelo_importacao_colaboradores.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.markdown("**Colunas recomendadas:**")
        st.code("matricula | nome | jornada_trabalho | cargo | setor | logins_jms | gestor_responsavel | folga_dominical | genero | ativo")

        arquivo = st.file_uploader(
            t("Arquivo Excel"),
            type=["xlsx"]
        )

        if arquivo is not None:
            try:
                excel = pd.ExcelFile(arquivo)

                col_imp1, col_imp2 = st.columns(2)

                with col_imp1:
                    aba_excel = st.selectbox(t("Aba da planilha"), excel.sheet_names)

                with col_imp2:
                    linha_cabecalho = st.number_input(
                        t("Linha onde está o cabeçalho"),
                        min_value=1,
                        value=1,
                        step=1
                    )

                arquivo.seek(0)

                df_importado = pd.read_excel(
                    arquivo,
                    sheet_name=aba_excel,
                    header=int(linha_cabecalho) - 1
                )

                df_preparado = preparar_dataframe_importacao(df_importado)
                df_analisado = analisar_duplicidades_importacao(df_preparado)

                total_novos = len(df_analisado[df_analisado["situacao_importacao"].astype(str).str.startswith("Novo")])
                total_atualizacoes = len(df_analisado[df_analisado["situacao_importacao"].astype(str).str.contains("Atualização", na=False)])
                total_duplicados_arquivo = len(df_analisado[df_analisado["situacao_importacao"].astype(str).str.contains("Duplicado no arquivo", na=False)])

                colm1, colm2, colm3 = st.columns(3)
                colm1.metric("Novos registros", total_novos)
                colm2.metric("Atualizações identificadas", total_atualizacoes)
                colm3.metric("Duplicatas no arquivo", total_duplicados_arquivo)

                st.write("Revise, corrija, exclua ou acrescente linhas antes de importar:")

                df_editado = st.data_editor(
                    df_analisado,
                    use_container_width=True,
                    num_rows="dynamic",
                    hide_index=True,
                    key="editor_importacao_colaboradores",
                    column_config={
                        "situacao_importacao": st.column_config.TextColumn(
                            "Situação da importação",
                            disabled=True
                        ),
                        "id_existente": st.column_config.TextColumn(
                            "ID existente",
                            disabled=True
                        ),
                        "folga_dominical": st.column_config.SelectboxColumn(t("Folga Dominical"), options=["Sim", "Não"]),
                        "genero": st.column_config.SelectboxColumn(t("Gênero"), options=["", "Feminino", "Masculino", "Outro"]),
                        "ativo": st.column_config.CheckboxColumn("Ativo")
                    }
                )

                atualizar_existentes = st.checkbox(
                    "Atualizar colaboradores existentes por matrícula ou nome",
                    value=True
                )

                atualizacao_incremental = st.checkbox(
                    "Atualização incremental: preencher somente campos novos ou informados sem apagar campos vazios",
                    value=True
                )

                if st.button("Importar dados para o sistema"):
                    resultado = importar_colaboradores(
                        df_editado,
                        atualizar_existentes=atualizar_existentes,
                        atualizacao_incremental=atualizacao_incremental
                    )

                    st.success(
                        f"Importação concluída. "
                        f"Inseridos: {resultado['inseridos']} | "
                        f"Atualizados: {resultado['atualizados']} | "
                        f"Ignorados: {resultado['ignorados']} | "
                        f"Duplicatas sinalizadas no arquivo: {resultado['duplicados_arquivo']}"
                    )

            except Exception as erro:
                st.error(f"Erro ao importar planilha: {erro}")

    with aba3:
        st.subheader(t("Cadastro e alteração de colaboradores"))
        st.caption(f"Filial padrão: {FILIAL_PADRAO}")

        sub1, sub2 = st.tabs([
            "Cadastrar novo",
            "Alterar cadastro"
        ])

        with sub1:
            with st.form("form_colaborador"):
                col1, col2 = st.columns(2)

                with col1:
                    matricula = st.text_input(t("Matrícula"))
                    nome = st.text_input(t("Nome"))
                    jornada_trabalho = st.text_input(t("Jornada de Trabalho"))
                    cargo = st.text_input(t("Cargo"))

                with col2:
                    setor = st.text_input(t("Setor"))
                    logins_jms = st.selectbox("Logins - JMS", ["Sim", "Não Precisa", "Não"])
                    gestor_responsavel = st.text_input(t("Responsável"))
                    folga_dominical = st.selectbox(t("Folga Dominical"), ["Sim", "Não"])
                    genero = st.selectbox(t("Gênero"), ["", "Feminino", "Masculino", "Outro"])
                    ativo = st.checkbox("Ativo", value=True)

                salvar = st.form_submit_button("Cadastrar")

                if salvar:
                    if not nome.strip():
                        st.warning("Informe o nome do colaborador.")
                    else:
                        cadastrar_colaborador(
                            matricula,
                            nome,
                            jornada_trabalho,
                            cargo,
                            setor,
                            logins_jms,
                            gestor_responsavel,
                            folga_dominical,
                            genero,
                            ativo
                        )
                        st.success("Colaborador cadastrado com sucesso.")

        with sub2:
            df_colaboradores = listar_colaboradores(ativos=False)

            if df_colaboradores.empty:
                st.warning("Nenhum colaborador cadastrado.")
            else:
                pesquisa = st.text_input(
                    "Pesquisar por nome, matrícula, cargo, setor ou responsável",
                    key="pesquisa_alteracao_cadastro"
                )

                df_resultado = filtrar_dataframe_colaborador(df_colaboradores, pesquisa)

                if df_resultado.empty:
                    st.warning("Nenhum registro localizado.")
                else:
                    df_resultado = df_resultado.copy()
                    df_resultado["opcao"] = (
                        df_resultado["matricula"].astype(str) +
                        " | " +
                        df_resultado["nome"].astype(str) +
                        " | " +
                        df_resultado["gestor_responsavel"].astype(str)
                    )

                    opcao = st.selectbox(
                        "Selecione o colaborador para alteração",
                        df_resultado["opcao"].tolist()
                    )

                    registro = df_resultado[df_resultado["opcao"] == opcao].iloc[0]

                    with st.form("form_alterar_colaborador"):
                        col1, col2 = st.columns(2)

                        with col1:
                            nova_matricula = st.text_input(
                                t("Matrícula"),
                                value=str(registro["matricula"])
                            )
                            novo_nome = st.text_input(
                                t("Nome"),
                                value=str(registro["nome"])
                            )
                            nova_jornada = st.text_input(
                                t("Jornada de Trabalho"),
                                value=str(registro["jornada_trabalho"])
                            )
                            novo_cargo = st.text_input(
                                t("Cargo"),
                                value=str(registro["cargo"])
                            )

                        with col2:
                            novo_setor = st.text_input(
                                t("Setor"),
                                value=str(registro["setor"])
                            )
                            novo_login = st.selectbox(
                                "Logins - JMS",
                                ["Sim", "Não Precisa", "Não"],
                                index=["Sim", "Não Precisa", "Não"].index(str(registro["logins_jms"]))
                                if str(registro["logins_jms"]) in ["Sim", "Não Precisa", "Não"]
                                else 0
                            )
                            novo_responsavel = st.text_input(
                                t("Responsável"),
                                value=str(registro["gestor_responsavel"])
                            )
                            novo_folga_dominical = st.selectbox(
                                t("Folga Dominical"),
                                ["Sim", "Não"],
                                index=["Sim", "Não"].index(str(registro.get("folga_dominical", "Sim")))
                                if str(registro.get("folga_dominical", "Sim")) in ["Sim", "Não"]
                                else 0
                            )
                            novo_genero = st.selectbox(
                                t("Gênero"),
                                ["", "Feminino", "Masculino", "Outro"],
                                index=["", "Feminino", "Masculino", "Outro"].index(str(registro.get("genero", "")))
                                if str(registro.get("genero", "")) in ["", "Feminino", "Masculino", "Outro"]
                                else 0
                            )
                            novo_ativo = st.checkbox(
                                "Ativo",
                                value=bool(int(registro["ativo"])) if str(registro["ativo"]).isdigit() else bool(registro["ativo"])
                            )

                        salvar_alteracao = st.form_submit_button("Salvar alteração do cadastro")

                        if salvar_alteracao:
                            if not novo_nome.strip():
                                st.warning("Informe o nome do colaborador.")
                            else:
                                atualizar_colaborador(
                                    registro["id"],
                                    nova_matricula,
                                    novo_nome,
                                    nova_jornada,
                                    novo_cargo,
                                    novo_setor,
                                    novo_login,
                                    novo_responsavel,
                                    novo_folga_dominical,
                                    novo_genero,
                                    novo_ativo
                                )
                                st.success("Cadastro alterado com sucesso.")

    with aba4:
        st.subheader(t("Base de colaboradores"))
        st.caption("Edite os dados e clique em salvar alterações. A filial permanece MG CGE.")

        df_colaboradores = listar_colaboradores(ativos=False)

        if df_colaboradores.empty:
            st.warning("Nenhum colaborador cadastrado.")
        else:
            pesquisa_base = st.text_input(
                "Pesquisar na base de colaboradores",
                key="pesquisa_base_colaboradores"
            )

            df_colaboradores = filtrar_dataframe_colaborador(df_colaboradores, pesquisa_base)

            if df_colaboradores.empty:
                st.warning("Nenhum colaborador localizado.")
            else:
                df_colaboradores["ativo"] = df_colaboradores["ativo"].apply(
                    lambda x: bool(int(x)) if str(x).isdigit() else bool(x)
                )

                df_editado = st.data_editor(
                    df_colaboradores,
                    use_container_width=True,
                    num_rows="dynamic",
                    hide_index=True,
                    key="editor_colaboradores",
                    column_config={
                        "id": st.column_config.NumberColumn("ID", disabled=True),
                        "filial": st.column_config.TextColumn("Filial", disabled=True),
                        "folga_dominical": st.column_config.SelectboxColumn(t("Folga Dominical"), options=["Sim", "Não"]),
                        "genero": st.column_config.SelectboxColumn(t("Gênero"), options=["", "Feminino", "Masculino", "Outro"]),
                        "ativo": st.column_config.CheckboxColumn("Ativo")
                    }
                )

                if st.button("Salvar alterações da base"):
                    resultado = salvar_edicao_colaboradores(df_editado)

                    st.success(
                        f"Base atualizada. "
                        f"Atualizados: {resultado['atualizados']} | "
                        f"Inseridos: {resultado['inseridos']} | "
                        f"Ignorados: {resultado['ignorados']}"
                    )

    with aba5:
        pagina_escala_folga_dominical(modo="Gestor")

    with aba6:
        st.subheader(t("Exportar Excel"))

        df = carregar_presencas()
        colaboradores = listar_colaboradores(ativos=True)

        if df.empty:
            st.warning("Não existem dados para exportar.")
        else:
            col1, col2 = st.columns(2)

            with col1:
                data_exportacao = st.date_input("Data para exportação", value=date.today())

            with col2:
                responsavel = st.selectbox(
                    t("Responsável"),
                    opcoes_unicas(colaboradores, "gestor_responsavel", "Todos"),
                    key="export_responsavel"
                )

            df["data_dt"] = pd.to_datetime(df["data"], errors="coerce").dt.date
            df_exportacao = df[df["data_dt"] == data_exportacao].copy()

            if responsavel != "Todos":
                df_exportacao = df_exportacao[df_exportacao["gestor_responsavel"] == responsavel]

            if df_exportacao.empty:
                st.warning("Não existem lançamentos para a data e responsável selecionados.")
            else:
                st.write("Revise os dados antes de baixar:")

                df_editado = st.data_editor(
                    df_exportacao,
                    use_container_width=True,
                    hide_index=True,
                    key="editor_exportacao"
                )

                arquivo_excel = gerar_excel_modelo(
                    df_editado,
                    data_exportacao,
                    responsavel
                )

                st.download_button(
                    label="Baixar lista de presença RH",
                    data=arquivo_excel,
                    file_name=f"lista_presenca_rh_{data_exportacao}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    with aba7:
        st.subheader(t("Gestão de acessos"))

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**Criar novo acesso**")

            with st.form("form_criar_acesso"):
                novo_usuario = st.text_input(t("Usuário"))
                nova_senha = st.text_input(t("Senha"), type="password")
                novo_perfil = st.selectbox("Perfil", PERFIS_USUARIO)

                criar_acesso = st.form_submit_button("Criar acesso")

                if criar_acesso:
                    resultado = criar_usuario(
                        novo_usuario,
                        nova_senha,
                        novo_perfil
                    )

                    if resultado["sucesso"]:
                        st.success(resultado["mensagem"])
                    else:
                        st.warning(resultado["mensagem"])

        with col2:
            st.markdown("**Alterar acesso existente**")

            df_usuarios = listar_usuarios()

            if df_usuarios.empty:
                st.warning("Nenhum usuário cadastrado.")
            else:
                usuario_opcao = st.selectbox(
                    "Usuário",
                    df_usuarios["usuario"].tolist()
                )

                registro_usuario = df_usuarios[df_usuarios["usuario"] == usuario_opcao].iloc[0]

                with st.form("form_alterar_acesso"):
                    perfil_atual = str(registro_usuario["perfil"])
                    perfil_index = PERFIS_USUARIO.index(perfil_atual) if perfil_atual in PERFIS_USUARIO else 0

                    perfil_editado = st.selectbox(
                        "Perfil",
                        PERFIS_USUARIO,
                        index=perfil_index
                    )

                    ativo_editado = st.checkbox(
                        "Ativo",
                        value=bool(registro_usuario["ativo"])
                    )

                    senha_editada = st.text_input(
                        "Nova senha",
                        type="password",
                        placeholder="Preencha somente se quiser trocar a senha"
                    )

                    salvar_acesso = st.form_submit_button("Salvar alteração do acesso")

                    if salvar_acesso:
                        atualizar_usuario(
                            registro_usuario["id"],
                            perfil_editado,
                            ativo_editado,
                            senha_editada if senha_editada else None
                        )

                        st.success("Acesso atualizado com sucesso.")

        st.divider()

        st.markdown("**Usuários cadastrados**")

        df_usuarios = listar_usuarios()

        if df_usuarios.empty:
            st.warning("Nenhum usuário cadastrado.")
        else:
            st.dataframe(
                df_usuarios,
                use_container_width=True,
                hide_index=True
            )


# =========================
# INICIALIZAÇÃO SEGURA DO BANCO
# =========================

def inicializar_banco():
    try:
        criar_tabelas()
    except psycopg2.Error as erro:
        st.error("Conexão estabelecida, mas houve erro ao criar ou validar as tabelas no Supabase.")
        st.caption("Verifique se o usuário do banco tem permissão de CREATE/ALTER no schema public.")
        st.code(str(erro).strip(), language="text")
        st.stop()
    except Exception as erro:
        st.error("Erro inesperado ao inicializar o banco de dados.")
        st.code(str(erro).strip(), language="text")
        st.stop()


# =========================
# APP PRINCIPAL
# =========================

inicializar_banco()

if "logado" not in st.session_state:
    st.session_state["logado"] = False

if not st.session_state["logado"]:
    login()
else:
    seletor_idioma("idioma_sidebar")
    st.sidebar.write(f"{t("Usuário")}: **{st.session_state['usuario']}**")
    st.sidebar.write(f"{t("Perfil")}: **{st.session_state['perfil']}**")
    st.sidebar.write(f"{t("Filial")}: **{FILIAL_PADRAO}**")

    if st.sidebar.button(t("Sair")):
        st.session_state.clear()
        st.rerun()

    if st.session_state["perfil"] == "Gestor":
        pagina_gestor()

    elif st.session_state["perfil"] == "Operação":
        pagina_operacao()